"""
一次性导入：把现有 Excel 的消费记录转成 import.sql，
按下单当天的历史中间汇率算出 MYR，再粘到 Supabase SQL Editor 执行。

用法：
    python scripts/import_excel.py <你的-supabase-user-id> [Excel路径]

- user-id：在 Supabase 注册并登录后，去 Authentication → Users 里复制你账号的 UID。
  也可以先不填，生成的 SQL 里会留占位符 REPLACE_WITH_YOUR_USER_ID，自己替换也行。
- Excel路径：默认 C:\\Users\\jiesh\\Downloads\\In game purchase.xlsx

依赖：pip install -r scripts/requirements.txt
"""

import sys
import time
import datetime as dt
from pathlib import Path

import requests
from openpyxl import load_workbook

DEFAULT_XLSX = r"C:\Users\jiesh\Downloads\In game purchase.xlsx"
OUT_SQL = Path(__file__).with_name("import.sql")
BASE = "myr"  # 基准货币（小写）

# Excel 里的币种写法 → 标准 ISO 代码
CURRENCY_MAP = {
    "JP¥": "JPY",
    "JPY": "JPY",
    "¥": "JPY",
    "TWD": "TWD",
    "NT$": "TWD",
    "KRW": "KRW",
    "₩": "KRW",
    "MYR": "MYR",
    "RM": "MYR",
}

_rate_cache: dict[tuple[str, str], float | None] = {}


def fetch_rate_on(date: str, base: str) -> float | None:
    """1 单位 base = ? MYR；取不到返回 None。"""
    b = base.lower()
    urls = [
        f"https://cdn.jsdelivr.net/npm/@fawazahmed0/currency-api@{date}/v1/currencies/{b}.json",
        f"https://{date}.currency-api.pages.dev/v1/currencies/{b}.json",
    ]
    for url in urls:
        try:
            r = requests.get(url, timeout=15)
            if r.status_code != 200:
                continue
            data = r.json()
            rate = data.get(b, {}).get(BASE)
            if isinstance(rate, (int, float)) and rate > 0:
                return float(rate)
        except Exception:
            continue
    return None


def get_rate(currency: str, order_date: str) -> tuple[float | None, str | None]:
    """返回 (汇率, 实际取到汇率的日期)。MYR 直接 1。向前回退最多 7 天。"""
    if currency == "MYR":
        return 1.0, order_date
    key = (currency, order_date)
    if key in _rate_cache:
        cached = _rate_cache[key]
        return cached, order_date if cached is not None else None
    base_date = dt.date.fromisoformat(order_date)
    for i in range(8):
        d = (base_date - dt.timedelta(days=i)).isoformat()
        rate = fetch_rate_on(d, currency)
        if rate is not None:
            _rate_cache[key] = rate
            return rate, d
        time.sleep(0.1)
    _rate_cache[key] = None
    return None, None


def sql_str(v) -> str:
    if v is None or v == "":
        return "null"
    return "'" + str(v).strip().replace("'", "''") + "'"


def sql_num(v) -> str:
    return "null" if v is None else repr(round(float(v), 6))


def main() -> None:
    user_id = sys.argv[1] if len(sys.argv) > 1 else "REPLACE_WITH_YOUR_USER_ID"
    xlsx = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(DEFAULT_XLSX)
    if not xlsx.exists():
        sys.exit(f"找不到 Excel：{xlsx}")

    wb = load_workbook(xlsx, data_only=True)
    ws = wb["Main"] if "Main" in wb.sheetnames else wb.worksheets[0]

    rows = []
    skipped = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        order_date, game, product, currency, cost = r[0], r[1], r[2], r[3], r[4]
        status = r[5] if len(r) > 5 else None
        if order_date is None or cost is None:
            continue
        if isinstance(order_date, dt.datetime):
            order_date = order_date.date()
        order_date = order_date.isoformat()
        cur = CURRENCY_MAP.get(str(currency).strip(), str(currency).strip().upper())

        rate, rate_date = get_rate(cur, order_date)
        if rate is None:
            print(f"  ⚠ 取不到汇率：{order_date} {cur} {cost} —— 该行 MYR 先记 0，导入后手动改")
            myr = 0.0
            skipped += 1
        else:
            myr = round(float(cost) * rate, 2)
        note = None
        if rate_date and rate_date != order_date:
            note = f"汇率用 {rate_date}"
        print(f"  {order_date}  {game:<20} {cur} {cost} -> RM {myr}")
        rows.append(
            dict(
                order_date=order_date,
                game=game,
                product=product,
                currency=cur,
                cost=cost,
                status=status,
                rate=rate,
                myr=myr,
                note=note,
            )
        )

    lines = [
        "-- 由 import_excel.py 生成。粘到 Supabase → SQL Editor → Run。",
        "-- 若顶部是 REPLACE_WITH_YOUR_USER_ID，请先全部替换成你的 user id。",
        "insert into public.purchases",
        "  (user_id, order_date, game, product_name, currency, cost, status, rate, rate_source, myr, note)",
        "values",
    ]
    values = []
    for x in rows:
        values.append(
            "  ("
            + ", ".join(
                [
                    sql_str(user_id),
                    sql_str(x["order_date"]),
                    sql_str(x["game"]),
                    sql_str(x["product"]),
                    sql_str(x["currency"]),
                    sql_num(x["cost"]),
                    sql_str(x["status"]),
                    sql_num(x["rate"]),
                    "'auto'",
                    sql_num(x["myr"]),
                    sql_str(x["note"]),
                ]
            )
            + ")"
        )
    lines.append(",\n".join(values) + ";")

    OUT_SQL.write_text("\n".join(lines), encoding="utf-8")
    print(f"\n✅ 已写出 {len(rows)} 条 → {OUT_SQL}")
    if skipped:
        print(f"   其中 {skipped} 条没取到汇率（MYR 记 0），导入后在网站里手动补。")
    if user_id == "REPLACE_WITH_YOUR_USER_ID":
        print("   ⚠ 记得把 SQL 里的 REPLACE_WITH_YOUR_USER_ID 换成你的 user id 再执行。")


if __name__ == "__main__":
    main()
